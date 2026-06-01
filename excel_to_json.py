#!/usr/bin/env python3
"""
統覧 TORAN — Excel → JSON 変換スクリプト
==========================================
使い方:
  python excel_to_json.py  （同フォルダのExcelファイルを自動検出）
  python excel_to_json.py --excel 名前リスト.xlsx

出力:
  output/politicians_base.json  ← 全813名の基本情報
  output/summary.txt            ← 変換サマリー
"""

import openpyxl, json, re, argparse, os, sys
from datetime import date

EXCEL_PATH_DEFAULT = '/root/.claude/uploads/efa3b98c-ee93-4e72-84d6-76a0936ba78e/af27dee2-________.xlsx'
OUTPUT_DIR = 'output'

# ── 会派→正式政党名マッピング ─────────────────────────
PARTY_MAP = {
    '自民':     '自民党',
    '自由民主': '自民党',
    '公明':     '公明党',
    '立憲':     '立憲民主党',
    '維新':     '日本維新の会',
    '国民':     '国民民主党',
    '民主':     '国民民主党',   # 参院: 国民民主系
    '共産':     '日本共産党',
    'れいわ':   'れいわ新選組',
    'れ新':     'れいわ新選組',
    '参政':     '参政党',
    '有志':     '有志の会',
    '社民':     '社民党',
    '無所属':   '無所属',
    '無':       '無所属',
    'みらい':   'チームみらい',  # 安野貴博代表
    'みら':     'チームみらい',
    '中道':     '（要確認）',    # 公明+立憲混在 → 手動で振り分け必要
    '保守':     '日本保守党',    # 百田尚樹・北村晴男ら
    '沖縄':     '沖縄社会大衆党',
    '教育':     '日本教育党',
    'NHK':      'NHK党',
    '政女':     '政女党',
}

def normalize_party(raw: str) -> str:
    if not raw:
        return '無所属'
    raw = str(raw).strip()
    for key, val in PARTY_MAP.items():
        if key in raw:
            return val
    return raw  # マッピングにない場合はそのまま

# ── 名前クリーニング ──────────────────────────────────
def clean_name(raw) -> str:
    if not raw:
        return ''
    s = str(raw)
    s = re.sub(r'[　 ]+', ' ', s)   # 全角・半角スペースを半角1つに
    s = s.rstrip('君').strip()        # 末尾の「君」を削除
    return s

def clean_reading(raw) -> str:
    if not raw:
        return ''
    s = str(raw)
    s = re.sub(r'[　 ]+', ' ', s)    # スペース正規化
    s = re.sub(r'[ァ-ン]', lambda m:   # カタカナ→ひらがな
        chr(ord(m.group(0)) - 0x60), s)
    s = s.strip().lower()
    return s

def clean_district(raw) -> str:
    if not raw:
        return ''
    s = str(raw).strip()
    # 選挙区を整形: （比）北関東 → 比例・北関東
    s = re.sub(r'[（(]比[）)]', '比例・', s)
    return s

# ── 年齢計算 ─────────────────────────────────────────
def calc_age_from_str(raw) -> int | None:
    if not raw:
        return None
    s = str(raw)
    # 「昭和56年」など→ 西暦変換
    m = re.search(r'昭和\s*(\d+)', s)
    if m:
        year = 1925 + int(m.group(1))
        return date.today().year - year
    m = re.search(r'平成\s*(\d+)', s)
    if m:
        year = 1988 + int(m.group(1))
        return date.today().year - year
    m = re.search(r'令和\s*(\d+)', s)
    if m:
        year = 2018 + int(m.group(1))
        return date.today().year - year
    m = re.search(r'(\d{4})', s)
    if m:
        return date.today().year - int(m.group(1))
    return None

# ── 名前から性別推定（簡易）────────────────────────────
FEMALE_CHARS = set('子美恵絵江花香菜奈愛里佳代和夏春恵智美由')
FEMALE_ENDS  = ['子','美','恵','絵','江','香','奈','愛','佳','代','夏','春','智']

def guess_gender(name: str, reading: str) -> str:
    """名前末尾と読みから性別を簡易推定。不明確は「男」（後で手動修正）"""
    # 名部分を取り出す（スペース後）
    parts = name.split()
    given = parts[-1] if parts else name
    if given:
        last_char = given[-1]
        if last_char in FEMALE_CHARS:
            return '女'
        for end in FEMALE_ENDS:
            if given.endswith(end):
                return '女'
    return '男'

# ── 空テンプレートエントリ ───────────────────────────
def make_entry(pid: str, name: str, reading: str, party: str,
               chamber: str, district: str, age: int | None,
               gender: str, status: str = '現職', faction: str = '') -> dict:
    return {
        "id":       pid,
        "name":     name,
        "reading":  reading,
        "party":    party,
        "faction":  faction,   # Excelの元会派名（確認用）
        "role":     "（要確認）",
        "chamber":  chamber,
        "district": district,
        "status":   status,
        "gender":   gender,
        "age":      age,
        # ── スコア（AI入力待ち）──
        "total":    0,
        "rank":     "未評価",
        "axes":     [0, 0, 0, 0, 0, 0, 0, 0],
        "stances":  {k: "△" for k in [
            "tax_cut","active_fiscal","discipline","defense","econ_sec",
            "immigration","renewable","nuclear","expo","ir","mynumber",
            "birthrate","education","regional","china","foreign","food","semi"
        ]},
        "plus":        "（AI評価待ち）",
        "minus":       "（AI評価待ち）",
        "comment":     "（AI評価待ち）",
        "flag_crime":  False,
        "flag_caution": False,
        "updated":     date.today().strftime("%Y.%m"),
        "survey":      "基本情報収集済",
    }

# ── シート別変換 ─────────────────────────────────────
def parse_sheet_giin(ws, chamber: str, id_offset: int) -> list[dict]:
    """衆議院・参議院シートを変換"""
    # ヘッダー行を確認
    # 衆院: A=No, B=区分, C=氏名, D=ふりがな, E=会派, F=選挙区
    # 参院: A=No, B=区分, C=議員氏名, D=読み方, E=会派, F=選挙区
    results = []
    idx = id_offset
    for row in range(2, ws.max_row + 1):
        name_raw    = ws.cell(row, 3).value
        reading_raw = ws.cell(row, 4).value
        party_raw   = ws.cell(row, 5).value
        district_raw= ws.cell(row, 6).value

        if not name_raw:
            continue

        name     = clean_name(name_raw)
        reading  = clean_reading(reading_raw)
        party    = normalize_party(str(party_raw) if party_raw else '')
        district = clean_district(district_raw)
        gender   = guess_gender(name, reading)
        pid      = f"P{idx:03d}"

        results.append(make_entry(pid, name, reading, party, chamber, district, None, gender,
                                   faction=str(party_raw).strip() if party_raw else ''))
        idx += 1

    return results

def parse_sheet_chiji(ws, id_offset: int) -> list[dict]:
    """知事シートを変換"""
    # A=No, B=都道府県名, C=知事名, D=ふりがな, E=生年月日, F=?, G=年齢
    results = []
    idx = id_offset
    for row in range(2, ws.max_row + 1):
        prefecture  = ws.cell(row, 2).value
        name_raw    = ws.cell(row, 3).value
        reading_raw = ws.cell(row, 4).value
        birth_raw   = ws.cell(row, 5).value
        age_raw     = ws.cell(row, 7).value

        if not name_raw:
            continue

        name     = clean_name(name_raw)
        reading  = clean_reading(reading_raw)
        district = str(prefecture).strip() if prefecture else ''
        gender   = guess_gender(name, reading)
        pid      = f"P{idx:03d}"

        # 年齢: G列優先、なければE列から計算
        age = None
        if age_raw and str(age_raw).isdigit():
            age = int(age_raw)
        else:
            age = calc_age_from_str(birth_raw)

        results.append(make_entry(pid, name, reading, '無所属', '首長', district, age, gender, faction='知事'))
        idx += 1

    return results

def parse_sheet_other(ws, chamber: str, id_offset: int, status: str = '現職') -> list[dict]:
    """議員候補・生存元議員シートを変換"""
    # A=No, B=区分, C=氏名, D=読み方, E=会派, F=生年月日, G=年齢
    results = []
    idx = id_offset
    for row in range(2, ws.max_row + 1):
        name_raw    = ws.cell(row, 3).value
        reading_raw = ws.cell(row, 4).value
        party_raw   = ws.cell(row, 5).value
        birth_raw   = ws.cell(row, 6).value
        age_raw     = ws.cell(row, 7).value

        if not name_raw:
            continue

        name     = clean_name(name_raw)
        reading  = clean_reading(reading_raw)
        party    = normalize_party(str(party_raw) if party_raw else '')
        gender   = guess_gender(name, reading)
        pid      = f"P{idx:03d}"

        age = None
        if age_raw and str(age_raw).strip().isdigit():
            age = int(age_raw)
        else:
            age = calc_age_from_str(birth_raw)

        results.append(make_entry(pid, name, reading, party, chamber, '', age, gender, status,
                                   faction=str(party_raw).strip() if party_raw else ''))
        idx += 1

    return results

# ── メイン ───────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--excel', default=EXCEL_PATH_DEFAULT)
    parser.add_argument('--out', default=f'{OUTPUT_DIR}/politicians_base.json')
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        sys.exit(f"❌ Excelファイルが見つかりません: {args.excel}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"📂 読み込み: {args.excel}\n")
    wb = openpyxl.load_workbook(args.excel)

    all_entries = []
    idx = 1  # P001 から連番

    # シート別に処理
    if '衆議院議員' in wb.sheetnames:
        entries = parse_sheet_giin(wb['衆議院議員'], '衆議院', idx)
        idx += len(entries)
        all_entries.extend(entries)
        print(f"✅ 衆議院議員: {len(entries)}名")

    if '参議院議員' in wb.sheetnames:
        entries = parse_sheet_giin(wb['参議院議員'], '参議院', idx)
        idx += len(entries)
        all_entries.extend(entries)
        print(f"✅ 参議院議員: {len(entries)}名")

    if '知事' in wb.sheetnames:
        entries = parse_sheet_chiji(wb['知事'], idx)
        idx += len(entries)
        all_entries.extend(entries)
        print(f"✅ 知事: {len(entries)}名")

    if '議員候補' in wb.sheetnames:
        entries = parse_sheet_other(wb['議員候補'], '候補', idx, status='候補')
        idx += len(entries)
        all_entries.extend(entries)
        print(f"✅ 議員候補: {len(entries)}名")

    if '生存元議員' in wb.sheetnames:
        entries = parse_sheet_other(wb['生存元議員'], '元議員', idx, status='元職')
        idx += len(entries)
        all_entries.extend(entries)
        print(f"✅ 生存元議員: {len(entries)}名")

    print(f"\n合計: {len(all_entries)}名\n")

    # 出力
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(all_entries, f, ensure_ascii=False, indent=2)
    print(f"📄 出力: {args.out}")

    # サマリー出力
    party_counts = {}
    gender_counts = {'男': 0, '女': 0}
    chamber_counts = {}
    for e in all_entries:
        party_counts[e['party']] = party_counts.get(e['party'], 0) + 1
        gender_counts[e['gender']] = gender_counts.get(e['gender'], 0) + 1
        chamber_counts[e['chamber']] = chamber_counts.get(e['chamber'], 0) + 1

    summary_path = f'{OUTPUT_DIR}/summary.txt'
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write(f"=== 統覧 TORAN — 変換サマリー ===\n")
        f.write(f"合計: {len(all_entries)}名\n\n")
        f.write("【院・区分別】\n")
        for k, v in sorted(chamber_counts.items(), key=lambda x: -x[1]):
            f.write(f"  {k}: {v}名\n")
        f.write("\n【政党別】\n")
        for k, v in sorted(party_counts.items(), key=lambda x: -x[1]):
            f.write(f"  {k}: {v}名\n")
        f.write("\n【性別】\n")
        for k, v in gender_counts.items():
            f.write(f"  {k}: {v}名\n")
        f.write("\n【次のステップ】\n")
        f.write("  1. output/politicians_base.json をAIに15名ずつ渡してスコア・根拠を追加\n")
        f.write("  2. 性別を手動確認（自動推定のため誤りがある場合あり）\n")
        f.write("  3. role（役職）を主要議員のみ手動入力\n")

    print(f"📋 サマリー: {summary_path}")

    # サマリーをコンソールにも表示
    print("\n【院・区分別】")
    for k, v in sorted(chamber_counts.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}名")
    print("\n【政党別 top10】")
    for k, v in sorted(party_counts.items(), key=lambda x: -x[1])[:10]:
        print(f"  {k}: {v}名")
    print("\n【性別（自動推定）】")
    for k, v in gender_counts.items():
        print(f"  {k}: {v}名")
    print("\n✅ 完了！")
    print("次: このJSONをAIに15名ずつ渡してスコア・根拠を追加してください。")

if __name__ == '__main__':
    main()
